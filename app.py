import os
import io
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import tempfile
import shutil
from datetime import datetime
from werkzeug.utils import secure_filename

def safe_read_excel(path, **kwargs):
    try:
        return pd.read_excel(path, engine='calamine', **kwargs)
    except Exception as e:
        if "engine" in str(e).lower() or "calamine" in str(e).lower():
            return pd.read_excel(path, **kwargs)
        raise e

def safe_excel_file(path, **kwargs):
    try:
        return pd.ExcelFile(path, engine='calamine', **kwargs)
    except Exception as e:
        if "engine" in str(e).lower() or "calamine" in str(e).lower():
            return pd.ExcelFile(path, **kwargs)
        raise e

import sys

if getattr(sys, 'frozen', False):
    # If the application is run as a bundle (PyInstaller)
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    # If run normally
    app = Flask(__name__)

app.secret_key = 'super_secret_key_change_me_in_production'

# Pastikan folder templates dan static ada
os.makedirs('templates', exist_ok=True)
os.makedirs('static', exist_ok=True)

COLUMN_ALIASES = {
    # Absensi & HKO Master
    'Pers.No.': ['Pers.No.', 'Personnel Number', 'NIK', 'KIT TK', 'Nomor Personil', 'No. Personil'],
    'Kode Mandor': ['Kode Mandor', 'Kit Mandor', 'Mandor', 'ID Mandor', 'Nomor Mandor'],
    'Full Name': ['Full Name', 'Nama', 'Nama Lengkap', 'Nama TK', 'Nama Pekerja'],
    'ZJABATAN': ['ZJABATAN', 'Jabatan', 'Position'],
    'Organizational Unit': ['Organizational Unit', 'Bagian', 'Unit', 'Departemen'],
    'Nama mandor': ['Nama mandor', 'Nama Mandor', 'Mandor Name'],
    
    # Absensi Report
    'KIT TK': ['KIT TK', 'Pers.No.', 'Personnel Number', 'NIK', 'Nomor Personil'],
    'Mandor': ['Mandor', 'Kode Mandor', 'Kit Mandor', 'ID Mandor'],
    'Nama TK': ['Nama TK', 'Full Name', 'Nama', 'Nama Lengkap'],
    'Date Shift': ['Date Shift', 'Date', 'Tanggal', 'Tgl', 'Tgl Shift'],
    
    # HKO
    'Start Date': ['Start Date', 'Date', 'Tanggal', 'Tgl', 'Tgl Mulai'],
    'Total Biaya': ['Total Biaya', 'Biaya', 'Cost', 'Total Cost', 'Total'],
    
    # Overtime
    'Personnel Number': ['Personnel Number', 'Pers.No.', 'NIK', 'KIT TK', 'Nomor Personil'],
    'Date': ['Date', 'Tanggal', 'Tgl', 'Start Date', 'Date Shift'],
    'Hari': ['Hari', 'Day'],
    'Bulan': ['Bulan', 'Month'],
    'Tahun': ['Tahun', 'Year'],
    'Absence Type': ['Absence Type', 'Tipe Absen', 'Jenis Absen', 'A/T'],
    'Jam Piket Biasa': ['Jam Piket Biasa', 'Piket Biasa', 'Biasa'],
    'Jam Piket Libur': ['Jam Piket Libur', 'Piket Libur', 'Libur']
}

def standardize_columns(df, expected_columns):
    """Mengubah nama kolom menjadi nama standar yang dibutuhkan program"""
    new_cols = {}
    for col in df.columns:
        col_str = str(col).strip()
        found = False
        for std_name in expected_columns:
            if std_name in COLUMN_ALIASES:
                if col_str.lower() in [a.lower() for a in COLUMN_ALIASES[std_name]]:
                    new_cols[col] = std_name
                    found = True
                    break
        if not found:
            new_cols[col] = col
    return df.rename(columns=new_cols)

def process_data_logic(master_path, report_path, output_path):
    # 1. Load Master Data
    master_df = safe_read_excel(master_path, sheet_name='Sheet1')
    master_df = standardize_columns(master_df, ['Pers.No.', 'Kode Mandor', 'Full Name', 'ZJABATAN', 'Organizational Unit', 'Nama mandor'])
    
    master_df['Pers.No.'] = master_df['Pers.No.'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    master_df['Kode Mandor'] = master_df['Kode Mandor'].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    name_lookup_pers = master_df.set_index('Pers.No.')['Full Name'].to_dict()
    jabatan_lookup = master_df.set_index('Pers.No.')['ZJABATAN'].to_dict()
    bagian_lookup = master_df.set_index('Pers.No.')['Organizational Unit'].to_dict()
    
    master_mandor = master_df[master_df['Kode Mandor'] != '']
    name_lookup_kode = master_mandor.set_index('Kode Mandor')['Nama mandor'].to_dict()

    # 2. Load Report Data
    xl_report = safe_excel_file(report_path)
    report_df = xl_report.parse('Worksheet')
    report_df = standardize_columns(report_df, ['KIT TK', 'Mandor', 'Nama TK', 'Date Shift'])
    
    report_df['KIT TK'] = report_df['KIT TK'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    report_df['Mandor'] = report_df['Mandor'].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    report_df['Nama TK'] = report_df['Nama TK'].astype(str).str.strip() # Menghapus spasi tersembunyi
    
    # 3. Hapus kolom jika sudah ada dari percobaan sebelumnya
    cols_to_drop = ['nama mandor', 'Jabatan', 'bagian', 'Totall']
    for col in cols_to_drop:
        if col in report_df.columns:
            report_df = report_df.drop(columns=[col])
            
    # 4. Tambahkan Kolom Baru
    mapped_by_pers = report_df['Mandor'].map(name_lookup_pers)
    mapped_by_kode = report_df['Mandor'].map(name_lookup_kode)
    nama_mandor_series = mapped_by_pers.fillna(mapped_by_kode)
    
    jabatan_series = report_df['KIT TK'].map(jabatan_lookup)
    bagian_series = report_df['KIT TK'].map(bagian_lookup)
    
    date_counts = report_df.groupby('KIT TK')['Date Shift'].nunique().to_dict()
    totall_series = report_df['KIT TK'].map(date_counts)
    
    col_mandor_idx = report_df.columns.get_loc('Mandor')
    report_df.insert(col_mandor_idx + 1, 'nama mandor', nama_mandor_series)
    report_df.insert(col_mandor_idx + 2, 'Jabatan', jabatan_series)
    report_df.insert(col_mandor_idx + 3, 'bagian', bagian_series)
    report_df.insert(col_mandor_idx + 4, 'Totall', totall_series)
    
    # Urutkan data berdasarkan Totall (Terbesar ke Terkecil) lalu berdasarkan Nama TK (A-Z)
    report_df = report_df.sort_values(by=['Totall', 'Nama TK'], ascending=[False, True])

    # Ganti NaNs dengan string kosong agar tidak error di openpyxl
    report_df = report_df.fillna("")

    # Gunakan file Template yang sudah mengandung PivotTable asli
    template_path = None
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        external_template = os.path.join(exe_dir, 'Template_Report.xlsx')
        if os.path.exists(external_template):
            template_path = external_template
        else:
            template_path = os.path.join(sys._MEIPASS, 'Template_Report.xlsx')
    else:
        template_path = os.path.join(app.root_path, 'Template_Report.xlsx')
        
    shutil.copy(template_path, output_path)

    # Buka file hasil kopian menggunakan openpyxl
    wb = openpyxl.load_workbook(output_path)
    
    if 'Worksheet' in wb.sheetnames:
        ws = wb['Worksheet']
        # Hapus data lama di Worksheet (sisakan header di baris 1)
        ws.delete_rows(2, ws.max_row)
        
        # Tulis data baru dari pandas DataFrame mulai baris 2
        for r_idx, row in enumerate(dataframe_to_rows(report_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
        # Format excel cells
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(color="000000", bold=True)
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                if cell.row == 1:
                    cell.fill = green_fill
                    cell.font = header_font
                    
        # Auto-adjust column widths roughly
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    wb.save(output_path)
    return output_path

def process_hko_logic(master_path, report_paths, output_path):
    dfs = []
    for path in report_paths:
        df = safe_read_excel(path, sheet_name='Sheet1')
        df = standardize_columns(df, ['Pers.No.', 'Start Date', 'Total Biaya'])
        
        if 'Pers.No.' not in df.columns or 'Start Date' not in df.columns or 'Total Biaya' not in df.columns:
            raise Exception("Kolom identitas (seperti 'Pers.No.', 'Start Date', atau 'Total Biaya') tidak ditemukan di salah satu file HKO. Pastikan format file benar.")
        # Jangan drop baris jika Start Date kosong, karena kita butuh Pers.No.-nya untuk Kehadiran 0
        df = df.dropna(subset=['Pers.No.'])
        
        # Ensure Start Date is datetime (coerce errors to NaT)
        df['Start Date'] = pd.to_datetime(df['Start Date'], errors='coerce')
        
        # Cari tanggal valid pertama di file ini untuk menebak bulan/tahun bagi yang Start Date-nya kosong
        valid_dates = df['Start Date'].dropna()
        if not valid_dates.empty:
            inferred_date = valid_dates.iloc[0]
            df['Start Date'] = df['Start Date'].fillna(inferred_date)
            
        dfs.append(df)
        
    combined_df = pd.concat(dfs, ignore_index=True)
    
    # 1. Jumlahkan Total Biaya per hari per orang (jika ada yang dobel di hari yang sama)
    daily_sum = combined_df.groupby(['Pers.No.', 'Start Date'])['Total Biaya'].sum().reset_index()
    
    # 1b. Dapatkan daftar semua (Pers.No, Bulan, Tahun) dari data mentah
    combined_df['Bulan'] = combined_df['Start Date'].dt.month
    combined_df['Tahun'] = combined_df['Start Date'].dt.year
    all_persons = combined_df[['Pers.No.', 'Bulan', 'Tahun']].drop_duplicates()
    
    # 2. Saring hanya hari yang Total Biayanya > 20
    valid_days = daily_sum[daily_sum['Total Biaya'] > 20].copy()
    
    # 3. Ekstrak Bulan dan Tahun
    valid_days['Bulan'] = valid_days['Start Date'].dt.month
    valid_days['Tahun'] = valid_days['Start Date'].dt.year
    
    # 4. Hitung Kehadiran (jumlah hari valid) per Bulan per Tahun
    valid_counts = valid_days.groupby(['Pers.No.', 'Bulan', 'Tahun']).size().reset_index(name='Kehadiran')
    
    # Gabungkan dengan all_persons agar yang 0 kehadiran tetap muncul
    result_df = pd.merge(all_persons, valid_counts, on=['Pers.No.', 'Bulan', 'Tahun'], how='left')
    result_df['Kehadiran'] = result_df['Kehadiran'].fillna(0).astype(int)
    
    # --- MAPPING MASTER DATA ---
    # Baca data master
    master_df = safe_read_excel(master_path, sheet_name='Sheet1')
    master_df = standardize_columns(master_df, ['Pers.No.', 'Kode Mandor', 'Full Name', 'Nama mandor'])
    
    master_df['Pers.No.'] = master_df['Pers.No.'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    master_df['Kode Mandor'] = master_df['Kode Mandor'].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    name_lookup_pers = master_df.set_index('Pers.No.')['Full Name'].to_dict()
    kode_mandor_lookup = master_df.set_index('Pers.No.')['Kode Mandor'].to_dict()
    
    master_mandor = master_df[master_df['Kode Mandor'] != '']
    name_lookup_kode = master_mandor.set_index('Kode Mandor')['Nama mandor'].to_dict()
    
    # Pastikan Pers.No. di result_df bertipe string untuk pencocokan
    result_df['Pers.No.'] = result_df['Pers.No.'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    # Map data
    result_df['Nama TK'] = result_df['Pers.No.'].map(name_lookup_pers).fillna('')
    result_df['Kode Mandor'] = result_df['Pers.No.'].map(kode_mandor_lookup).fillna('')
    
    # Untuk mendapatkan Nama Mandor, kita bisa menggunakan hasil 'Kode Mandor' yang baru saja di map
    # ATAU fallback ke lookup kode mandor via Pers.No. (seperti di absensi jika Mandor diisi No Pers)
    # Di sini, karena kita narik Kode Mandor dari Master (berdasarkan Pers.No TK), kita gunakan hasil tersebut
    result_df['Nama Mandor'] = result_df['Kode Mandor'].map(name_lookup_kode).fillna('')
    
    # Atur urutan kolom
    cols = ['Pers.No.', 'Bulan', 'Tahun', 'Kehadiran', 'Nama TK', 'Kode Mandor', 'Nama Mandor']
    result_df = result_df[cols]
    
    # 5. Urutkan berdasarkan Tahun, Bulan, lalu Pers.No. (Sesuai permintaan: bulan 4 kumpul dulu, baru bulan 5, dst)
    result_df = result_df.sort_values(by=['Tahun', 'Bulan', 'Pers.No.'])
    
    # Tulis hasil akhirnya ke Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='Rekap HKO', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Rekap HKO']
        
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(color="000000", bold=True)
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                if cell.row == 1:
                    cell.fill = green_fill
                    cell.font = header_font
                    
        for col in worksheet.columns:
            column = col[0].column_letter
            worksheet.column_dimensions[column].width = 15
            
    return output_path

def process_overtime_logic(input_path, output_path):
    # Read original data
    df = pd.read_excel(input_path, engine='calamine')
    
    # Standardize expected columns
    expected = ['Personnel Number', 'Date', 'Hari', 'Bulan', 'Tahun', 'Absence Type', 'Jam Piket Biasa', 'Jam Piket Libur']
    df = standardize_columns(df, expected)
    
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
    
    
    # Sheet 1: Data Absence
    cols_absence = ['Personnel Number', 'Date', 'Hari', 'Bulan', 'Tahun', 'Absence Type']
    available_absence = [c for c in cols_absence if c in df.columns]
    df_absence = df[available_absence].copy()
    
    # Sheet 2: Data Piket
    cols_piket = ['Personnel Number', 'Date', 'Hari', 'Bulan', 'Tahun', 'Jam Piket Biasa', 'Jam Piket Libur']
    available_piket = [c for c in cols_piket if c in df.columns]
    df_piket = df[available_piket].copy()
    
    # Add Total column for Piket
    if 'Jam Piket Biasa' in df_piket.columns and 'Jam Piket Libur' in df_piket.columns:
        df_piket['Total'] = pd.to_numeric(df_piket['Jam Piket Biasa'], errors='coerce').fillna(0) + pd.to_numeric(df_piket['Jam Piket Libur'], errors='coerce').fillna(0)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Original Data', index=False)
        df_absence.to_excel(writer, sheet_name='Data Absence', index=False)
        df_piket.to_excel(writer, sheet_name='Data Piket', index=False)
        
    # Apply styling using openpyxl
    wb = openpyxl.load_workbook(output_path)
    
    from openpyxl.styles import PatternFill, Font, Border, Side
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") # grey
    header_font = Font(color="003366", bold=True) # dark blue
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    for sheet_name in ['Data Absence', 'Data Piket']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Style headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                
            # Style data cells with border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
                    if ws.cell(row=1, column=cell.column).value == 'Date':
                        cell.number_format = 'yyyy-mm-dd'
                    
            # Auto-fit columns
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                
    wb.save(output_path)

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/process', methods=['POST'])
def process_files():
    if 'master_file' not in request.files or 'report_file' not in request.files:
        return "Missing files", 400
        
    master_file = request.files['master_file']
    report_file = request.files['report_file']
    
    if master_file.filename == '' or report_file.filename == '':
        return "No selected file", 400
        
    temp_dir = tempfile.mkdtemp()
    
    master_path = os.path.join(temp_dir, secure_filename(master_file.filename))
    report_path = os.path.join(temp_dir, secure_filename(report_file.filename))
    
    # Pastikan ekstensi selalu lowercase agar tidak error di pandas ExcelWriter
    base_name = secure_filename(report_file.filename).rsplit('.', 1)[0]
    output_filename = "Processed_" + base_name + ".xlsx"
    output_path = os.path.join(temp_dir, output_filename)
    
    master_file.save(master_path)
    report_file.save(report_path)
    
    try:
        process_data_logic(master_path, report_path, output_path)
        with open(output_path, 'rb') as f:
            data = f.read()
        return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=output_filename)
    except Exception as e:
        return str(e), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.route('/process_hko', methods=['POST'])
def process_hko():
    if 'hko_master_file' not in request.files or 'hko_file' not in request.files:
        return "Missing files", 400
        
    master_file = request.files['hko_master_file']
    hko_files = request.files.getlist('hko_file')
    
    if master_file.filename == '' or not hko_files or hko_files[0].filename == '':
        return "No selected file", 400
        
    temp_dir = tempfile.mkdtemp()
    
    master_path = os.path.join(temp_dir, secure_filename(master_file.filename))
    master_file.save(master_path)
    
    hko_paths = []
    for f in hko_files:
        path = os.path.join(temp_dir, secure_filename(f.filename))
        f.save(path)
        hko_paths.append(path)
    
    # Pastikan ekstensi selalu lowercase agar tidak error di pandas ExcelWriter
    base_name = secure_filename(hko_files[0].filename).rsplit('.', 1)[0]
    if len(hko_files) > 1:
        base_name += f"_dan_{len(hko_files)-1}_file_lainnya"
        
    output_filename = "Processed_HKO_" + base_name + ".xlsx"
    output_path = os.path.join(temp_dir, output_filename)
    
    try:
        process_hko_logic(master_path, hko_paths, output_path)
        with open(output_path, 'rb') as f:
            data = f.read()
        return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=output_filename)
    except Exception as e:
        return str(e), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.route('/process_overtime', methods=['POST'])
def process_overtime():
    if 'overtime_file' not in request.files:
        return "Missing file", 400
        
    overtime_file = request.files['overtime_file']
    if overtime_file.filename == '':
        return "No selected file", 400
        
    temp_dir = tempfile.mkdtemp()
    
    input_path = os.path.join(temp_dir, secure_filename(overtime_file.filename))
    
    base_name = secure_filename(overtime_file.filename).rsplit('.', 1)[0]
    output_filename = "Processed_" + base_name + ".xlsx"
    output_path = os.path.join(temp_dir, output_filename)
    
    overtime_file.save(input_path)
    
    try:
        process_overtime_logic(input_path, output_path)
        with open(output_path, 'rb') as f:
            data = f.read()
        return send_file(io.BytesIO(data), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=output_filename)
    except Exception as e:
        return str(e), 500
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

@app.route('/shutdown', methods=['POST'])
def shutdown():
    import threading
    import os
    import time
    
    def kill_server():
        time.sleep(1) # Beri waktu 1 detik agar browser sempat menerima balasan sukses
        os._exit(0)
        
    threading.Thread(target=kill_server, daemon=True).start()
    return jsonify({'message': 'Server telah dimatikan secara permanen. Anda sekarang bebas untuk memindahkan, mengganti nama, atau menghapus file Report_PG2.exe.'})

if __name__ == '__main__':
    from waitress import serve
    import threading
    import webbrowser
    import time
    
    def open_browser():
        time.sleep(1.5) # Beri waktu sebentar agar server Waitress nyala duluan
        webbrowser.open_new("http://127.0.0.1:5000")

    print("=========================================================")
    print(" Server Report Sistem PG2 (Production Mode) Aktif!")
    print(" Silakan buka browser di: http://127.0.0.1:5000")
    print("=========================================================")
    
    # Jalankan perintah pembuka browser di background
    threading.Thread(target=open_browser, daemon=True).start()
    
    # Jalankan server
    serve(app, host='127.0.0.1', port=5000, threads=8)
