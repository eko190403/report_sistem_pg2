import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

def process_overtime_logic(input_path, output_path):
    # Read original data
    df = pd.read_excel(input_path, engine='calamine')
    
    # Sheet 1: Data Absence
    cols_absence = ['Personnel Number', 'Date', 'Hari', 'Bulan', 'Tahun', 'Absence Type']
    df_absence = df[cols_absence].copy()
    
    # Sheet 2: Data Piket
    cols_piket = ['Personnel Number', 'Date', 'Hari', 'Bulan', 'Tahun', 'Jam Piket Biasa', 'Jam Piket Libur']
    df_piket = df[cols_piket].copy()
    
    # Add Total column for Piket
    df_piket['Total'] = df_piket['Jam Piket Biasa'].fillna(0) + df_piket['Jam Piket Libur'].fillna(0)
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Original Data', index=False)
        df_absence.to_excel(writer, sheet_name='Data Absence', index=False)
        df_piket.to_excel(writer, sheet_name='Data Piket', index=False)
        
    # Apply styling using openpyxl
    wb = openpyxl.load_workbook(output_path)
    
    # Define styles
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") # grey
    header_font = Font(color="003366", bold=True) # dark blue
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    for sheet_name in ['Data Absence', 'Data Piket']:
        ws = wb[sheet_name]
        
        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            
        # Style data cells with border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                
        # Auto-fit columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
            
    wb.save(output_path)

if __name__ == '__main__':
    process_overtime_logic('Overtime 7 2026.XLSX', 'scratch/test_overtime_out.xlsx')
    print("Done")
