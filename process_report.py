import pandas as pd
import os
import glob
from openpyxl.styles import PatternFill, Alignment, Font

def process_data(data_dir, master_file, report_file, output_file):
    print(f"Membaca data dari {data_dir}...")
    master_path = os.path.join(data_dir, master_file)
    report_path = os.path.join(data_dir, report_file)
    output_path = os.path.join(data_dir, output_file)

    master_df = pd.read_excel(master_path, sheet_name='Sheet1')
    master_df['Pers.No.'] = master_df['Pers.No.'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    master_df['Kode Mandor'] = master_df['Kode Mandor'].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    
    name_lookup_pers = master_df.set_index('Pers.No.')['Full Name'].to_dict()
    jabatan_lookup = master_df.set_index('Pers.No.')['ZJABATAN'].to_dict()
    bagian_lookup = master_df.set_index('Pers.No.')['Organizational Unit'].to_dict()
    
    master_mandor = master_df[master_df['Kode Mandor'] != '']
    name_lookup_kode = master_mandor.set_index('Kode Mandor')['Nama mandor'].to_dict()

    xl_report = pd.ExcelFile(report_path)
    report_df = xl_report.parse('Worksheet')
    
    report_df['KIT TK'] = report_df['KIT TK'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    report_df['Mandor'] = report_df['Mandor'].fillna('').astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    report_df['Nama TK'] = report_df['Nama TK'].astype(str).str.strip() # Menghapus spasi tersembunyi
    
    
    mapped_by_pers = report_df['Mandor'].map(name_lookup_pers)
    mapped_by_kode = report_df['Mandor'].map(name_lookup_kode)
    nama_mandor_series = mapped_by_pers.fillna(mapped_by_kode)
    
    jabatan_series = report_df['KIT TK'].map(jabatan_lookup)
    bagian_series = report_df['KIT TK'].map(bagian_lookup)
    
    date_counts = report_df.groupby('KIT TK')['Date Shift'].nunique().to_dict()
    totall_series = report_df['KIT TK'].map(date_counts)
    # Hapus kolom jika sudah ada dari percobaan sebelumnya
    cols_to_drop = ['nama mandor', 'Jabatan', 'bagian', 'Totall']
    for col in cols_to_drop:
        if col in report_df.columns:
            report_df = report_df.drop(columns=[col])
    
    col_mandor_idx = report_df.columns.get_loc('Mandor')
    report_df.insert(col_mandor_idx + 1, 'nama mandor', nama_mandor_series)
    report_df.insert(col_mandor_idx + 2, 'Jabatan', jabatan_series)
    report_df.insert(col_mandor_idx + 3, 'bagian', bagian_series)
    report_df.insert(col_mandor_idx + 4, 'Totall', totall_series)
    
    # Create the Pivot summary
    # Group by Nama TK, nama mandor, Jabatan, bagian, Date Shift
    # Sum the Totall
    # We will fill missing values with empty string for grouping if necessary, but we should just dropna=False
    # Wait, 'Date Shift' might be datetime. We can convert to string or keep as datetime.
    
    pivot_df = report_df.groupby(['Nama TK', 'nama mandor', 'Jabatan', 'bagian', 'Date Shift'])['Totall'].sum().reset_index()
    # Sort just like in the screenshot, but prioritized by Totall (largest to smallest)
    pivot_df = pivot_df.sort_values(
        by=['Totall', 'Nama TK', 'nama mandor', 'Jabatan', 'bagian', 'Date Shift'], 
        ascending=[False, True, True, True, True, True]
    )

    print(f"Menyimpan hasil ke {output_file}...")
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        report_df.to_excel(writer, sheet_name='Worksheet', index=False)
        pivot_df.to_excel(writer, sheet_name='Pivot Summary', index=False)
        
        for sheet_name in xl_report.sheet_names:
            if sheet_name != 'Worksheet' and sheet_name != 'Sheet3': # Skip old pivot
                other_df = xl_report.parse(sheet_name)
                other_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
        # Format excel cells
        workbook = writer.book
        green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(color="FFFFFF", bold=True)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = center_alignment
                    if cell.row == 1:
                        cell.fill = green_fill
                        cell.font = header_font
                        
            # Auto-adjust column widths roughly
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width
                
    print("Selesai!")

if __name__ == "__main__":
    data_dir = "."
    
    # Cari file master secara otomatis (biasanya yang namanya 12082026 atau sejenisnya)
    master_files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx') or f.endswith('.XLSX')]
    # Filter file yang bukan process result dan bukan BA plantation
    master_candidates = [f for f in master_files if 'Processed' not in f and 'BA plantation' not in f]
    report_candidates = [f for f in master_files if 'Processed' not in f and 'BA plantation' in f]
    
    master_file = master_candidates[0] if master_candidates else "12082026.XLSX"
    report_file = report_candidates[0] if report_candidates else "Report Data BA plantation2 - 01.08.2026 (3).xlsx"
    
    output_file = "Report Data BA plantation2 - Processed_Final2.xlsx"
    
    print(f"Master File: {master_file}")
    print(f"Report File: {report_file}")
    
    if os.path.exists(master_file) and os.path.exists(report_file):
        process_data(data_dir, master_file, report_file, output_file)
    else:
        print("File master atau report tidak ditemukan di direktori saat ini.")
